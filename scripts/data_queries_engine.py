import os, re, time
import pandas as pd
import numpy as np
from project_configs import PROJECT_CONFIGS, get_stock_thresholds


# ═══════════════════════════════════════════════════════════════
# ⚙️  CONFIG — change these each run
# ═══════════════════════════════════════════════════════════════
PROJECT_NAME = 'Kenya-MRA'  # Change this for each project run. Must match a key in PROJECT_CONFIGS.
DATA_PATH    = r"C:\Users\ID0373122\OneDrive - Kantar\Desktop\ALL PROJECTS\Kenya-MRA\Kenya-MRA-Data-Files\June2026\MRA_Kenya_Data_1 1.xlsx"
OUTPUT_DIR   = r"C:\Users\ID0373122\OneDrive - Kantar\Desktop\ALL PROJECTS\Kenya-MRA\Kenya-MRA-Outputs\June2026"
PREV_FEEDBACK_PATH = None

# Low-value suppression thresholds for stock/purchase outliers.
LOW_STOCK_QUERY_LIMIT    = 25
LOW_PURCHASE_QUERY_LIMIT = 35


# ═══════════════════════════════════════════════════════════════
# PART 1: COLUMN RESOLUTION
# ═══════════════════════════════════════════════════════════════

def normalize(name):
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

def deep_normalize(name):
    letters = re.sub(r'[^a-z]', ' ', str(name).lower())
    return re.sub(r'\s+', ' ', letters).strip()

def prefix_normalize(name):
    alnum = re.sub(r'[^a-z0-9]', ' ', str(name).lower())
    return re.sub(r'\s+', ' ', alnum).strip()

def resolve_col(df, target, must_not_contain=None):
    must_not_contain = must_not_contain or []
    if target in df.columns:
        return target
    tn = normalize(target)
    for col in df.columns:
        cn = normalize(col)
        if cn == tn:
            if not any(normalize(x) in cn for x in must_not_contain):
                return col
    return None

def resolve_all(df, config):
    resolved = {}
    for key in ['outlet_id','sku_id','channel','stock_col','purchase_col']:
        target = config.get(key)
        if target:
            found = resolve_col(df, target)
            resolved[key] = found
            if found is None:
                print(f"   WARNING: Could not find column '{target}' (key: {key})")
    resolved['price_cols'] = []
    for pc in config.get('price_cols', []):
        found = resolve_col(df, pc)
        if found:
            resolved['price_cols'].append(found)
        else:
            print(f"   WARNING: Price column '{pc}' not found")
    bp = config.get('buying_price_col')
    resolved['buying_price_col'] = resolve_col(df, bp) if bp else None
    uc = config.get('units_col')
    resolved['units_col'] = resolve_col(df, uc) if uc else None
    return resolved


# ═══════════════════════════════════════════════════════════════
# PART 2: DATA LOADING
# ═══════════════════════════════════════════════════════════════

def load_sheets(data_path):
    print(f"\n📂 Loading: {data_path}")
    xl = pd.ExcelFile(data_path)
    all_months = xl.sheet_names
    print(f"   Sheets: {all_months}")
    df_list = []
    for sheet in all_months:
        df = pd.read_excel(data_path, sheet_name=sheet)
        df.columns = df.columns.str.strip()
        df_list.append(df)
    return df_list, all_months, all_months[-1], all_months[-2], all_months[:-1]


# ═══════════════════════════════════════════════════════════════
# PART 3: UNITS RESOLUTION
# ═══════════════════════════════════════════════════════════════

def build_units_map(df_list, all_months, resolved):
    oid_col, sku_col, unt_col = (resolved.get('outlet_id'),
                                  resolved.get('sku_id'),
                                  resolved.get('units_col'))
    if not all([oid_col, sku_col, unt_col]):
        return {}
    rows = []
    for i, (sheet, df) in enumerate(zip(all_months, df_list)):
        if {oid_col, sku_col, unt_col}.issubset(df.columns):
            tmp = df[[oid_col, sku_col, unt_col]].copy().astype(str)
            tmp['_idx'] = i
            rows.append(tmp)
    if not rows:
        return {}
    all_u = pd.concat(rows, ignore_index=True)
    cnt_df = all_u.groupby([oid_col, sku_col, unt_col]).size().reset_index(name='n')
    curr_map = {(r[oid_col], r[sku_col]): r[unt_col]
                for _, r in all_u[all_u['_idx'] == len(all_months)-1].iterrows()}
    res = {}
    for (o, s), g in cnt_df.groupby([oid_col, sku_col]):
        mx   = g['n'].max()
        tied = sorted(g.loc[g['n']==mx, unt_col].tolist())
        cu   = curr_map.get((o, s))
        res[(o, s)] = cu if cu in tied else tied[0]
    return res

def add_comb_keys(df, resolved, config, units_map=None):
    oid, sku = resolved['outlet_id'], resolved['sku_id']
    unt = resolved.get('units_col')
    df = df.copy()
    df['Comb_stock'] = df[oid].astype(str) + df[sku].astype(str)
    if config.get('use_units_in_price_comb') and unt and unt in df.columns:
        if units_map:
            df['_res_unit'] = df.apply(
                lambda r: units_map.get((str(r[oid]), str(r[sku])), str(r[unt])), axis=1)
        else:
            df['_res_unit'] = df[unt].astype(str)
        df['Comb_price'] = df[oid].astype(str) + df[sku].astype(str) + df['_res_unit']
    else:
        df['Comb_price'] = df['Comb_stock']
    return df


# ═══════════════════════════════════════════════════════════════
# PART 4: SALES / NEGATIVE SALES
# ═══════════════════════════════════════════════════════════════

def calculate_sales(df_list, all_months, resolved, columns_to_check):
    oid_col, sku_col = resolved['outlet_id'], resolved['sku_id']
    stk_col, pch_col = resolved['stock_col'], resolved['purchase_col']
    dc, dp = df_list[-1].copy(), df_list[-2].copy()
    dc['Comb_stock'] = dc[oid_col].astype(str) + dc[sku_col].astype(str)
    dp['Comb_stock'] = dp[oid_col].astype(str) + dp[sku_col].astype(str)
    s = dc[['Comb_stock', pch_col, stk_col]].rename(
        columns={stk_col:'Current_Stock', pch_col:'Purchase'})
    p = dp[['Comb_stock', stk_col]].rename(columns={stk_col:'Previous_Stock'})
    s = s.merge(p, on='Comb_stock', how='left')
    s['Purchase'] = s['Purchase'].fillna(0)
    s['Current_Stock'] = s['Current_Stock'].fillna(0)
    s['Sales'] = np.where(s['Previous_Stock'].notna(),
                          s['Purchase'] + s['Previous_Stock'] - s['Current_Stock'],
                          np.nan)
    valid = s[s['Previous_Stock'].notna()].copy()
    sales_dict = valid.set_index('Comb_stock')['Sales'].to_dict()
    neg_rows = []
    for _, row in valid[valid['Sales'] < 0].iterrows():
        comb = row['Comb_stock']
        for df_m in reversed(df_list):
            df_m['Comb_stock'] = df_m[oid_col].astype(str) + df_m[sku_col].astype(str)
            match = df_m[df_m['Comb_stock'] == comb]
            if not match.empty:
                rec = match.iloc[0].to_dict()
                rec['Comb'] = rec['Comb_stock'] = str(comb)
                rec['Sales'] = row['Sales']
                rec['Queries'] = 'Negative sales'
                rec['Combination'] = 'Sales'
                for col in columns_to_check:
                    ac = resolve_col(df_m, col)
                    if ac:
                        for mo, dm in zip(all_months, df_list):
                            dm['Comb_stock'] = dm[oid_col].astype(str)+dm[sku_col].astype(str)
                            mm = dm[dm['Comb_stock']==comb]
                            if not mm.empty and ac in mm.columns:
                                rec[f"{col} ({mo})"] = mm.iloc[0][ac]
                neg_rows.append(rec)
                break
    print(f"   Negative Sales: {len(neg_rows)} flagged")
    return neg_rows, sales_dict


# ═══════════════════════════════════════════════════════════════
# PART 5A: PROFIT CHECKS
# ═══════════════════════════════════════════════════════════════

def check_profit(df_current, resolved, df_list, all_months, config):
    rows = []
    buy_col  = resolved.get('buying_price_col')
    sell_col = resolved['price_cols'][0] if resolved['price_cols'] else None
    oid_col, sku_col = resolved['outlet_id'], resolved['sku_id']
    if not buy_col or not sell_col: return rows
    if buy_col not in df_current.columns or sell_col not in df_current.columns: return rows
    df = df_current.copy()
    df['Comb_stock'] = df[oid_col].astype(str) + df[sku_col].astype(str)
    df[buy_col] = pd.to_numeric(df[buy_col], errors='coerce')
    df[sell_col] = pd.to_numeric(df[sell_col], errors='coerce')
    checks = []
    if 'negative_profit' in config.get('extra_checks', []):
        checks.append(('Negative Profit', df[buy_col] > df[sell_col]))
    if 'no_profit' in config.get('extra_checks', []):
        checks.append(('No Profit', df[buy_col] == df[sell_col]))
    for label, mask in checks:
        for _, row in df[mask].iterrows():
            rec = row.to_dict()
            rec['Comb_stock'] = str(row['Comb_stock'])
            rec['Comb_price'] = str(row['Comb_stock'])
            rec['Comb'] = str(row['Comb_stock'])
            rec['Queries'] = label
            rec['Combination'] = 'Profit Check'
            for pc in [sell_col, buy_col]:
                for mo, dm in zip(all_months, df_list):
                    mm = dm[dm[oid_col].astype(str)+dm[sku_col].astype(str)==row['Comb_stock']]
                    if not mm.empty and pc in mm.columns:
                        rec[f"{pc} ({mo})"] = mm.iloc[0][pc]
            rows.append(rec)
    print(f"   Profit checks: {len(rows)} flagged")
    return rows


# ═══════════════════════════════════════════════════════════════
# PART 5B: FACINGS CHECK
# ═══════════════════════════════════════════════════════════════

def find_facings_pairs(df):
    FACINGS_KWS  = {'facings','facing'}
    UNIT_KWS     = {'unit','units','count','total','stock'}
    IGNORE_WORDS = FACINGS_KWS | UNIT_KWS | {'no','the','and','for','of','a','an','front'}
    def _has(col, kws): return any(k in deep_normalize(col).split() for k in kws)
    def _prefix(col):
        return {w for w in prefix_normalize(col).split()
                if w not in IGNORE_WORDS and len(w) > 0}
    cols = list(df.columns)
    facings = [c for c in cols if _has(c, FACINGS_KWS)]
    units   = [c for c in cols if _has(c, UNIT_KWS) and not _has(c, FACINGS_KWS)]
    pairs = []
    for fc in facings:
        fp = _prefix(fc)
        best, best_s = None, -1
        for uc in units:
            s = len(fp & _prefix(uc)) if (fp or _prefix(uc)) else 0.5
            if s > best_s: best_s, best = s, uc
        if best:
            pairs.append((fc, best))
            print(f"   Facings pair: '{fc}' ↔ '{best}'")
    return pairs

def check_facings(df_current, resolved):
    pairs = find_facings_pairs(df_current)
    if not pairs: return []
    rows = []
    for fc, uc in pairs:
        df = df_current.copy()
        df[fc] = pd.to_numeric(df[fc], errors='coerce')
        df[uc] = pd.to_numeric(df[uc], errors='coerce')
        for _, row in df[df[fc] > df[uc]].iterrows():
            rec = row.to_dict()
            rec['Queries'] = 'Facings exceeds unit stock'
            rec['Combination'] = 'Facings'
            rows.append(rec)
    print(f"   Facings: {len(rows)} flagged")
    return rows


# ═══════════════════════════════════════════════════════════════
# PART 5C: FEEDBACK SUPPRESSION
# ═══════════════════════════════════════════════════════════════

_FA_KEYWORDS = [
    'correctly counted','correctly purchased','correctly recorded',
    'correct stock','stock correctly','purchase correctly',
    'confirmed correct','data is correct','data correct',
    'it is correct','is correct','are correct','was correct',
    'actually higher','actually lower','actually correct',
    'stocks only','due to low sales','low sales caused',
    'sold off','sold out','closing','shut down',
    'no money','no stock','out of stock',
    'restructur','seasonal','no error','ok','okay','valid','accurate',
]
_TO_KEYWORDS = [
    'error','wrong','mistake','incorrect','auditor',
    'should be','should have','correct value',
    'units purchased','units bought','price was','price is',
]

def _is_false_alarm_feedback(text):
    if pd.isna(text) or str(text).strip() == '': return False
    t = str(text).lower().strip()
    if any(kw in t for kw in _TO_KEYWORDS): return False
    return any(kw in t for kw in _FA_KEYWORDS)

def _find_feedback_col(columns):
    for c in columns:
        n = normalize(c)
        if n in ('feedback','feedbacks') or n.startswith('feedback'): return c
    return None

def _find_query_col(columns):
    for c in columns:
        n = normalize(c)
        if n in ('queries','query','querys'): return c
    for c in columns:
        n = normalize(c)
        if n.startswith('quer') and 'ask' not in n: return c
    return None

def _find_feedback_sheet(filepath):
    xl = pd.ExcelFile(filepath)
    if 'Data Queries' in xl.sheet_names: return 'Data Queries'
    for s in xl.sheet_names:
        if 'quer' in s.lower(): return s
    return xl.sheet_names[0]

def load_suppression_set(path, resolved):
    if not path: return {}
    if not os.path.exists(path):
        print(f"   WARNING: Feedback file not found: {path}")
        return {}
    print(f"\n🔁 Loading feedback: {path}")
    try:
        sheet = _find_feedback_sheet(path)
        try:
            df = pd.read_excel(path, sheet_name=sheet, header=1)
            df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            df = pd.read_excel(path, sheet_name=sheet, header=0)
            df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        print(f"   WARNING: Could not read: {e}")
        return {}
    oc = resolve_col(df, resolved.get('outlet_id','outletid'))
    sc = resolve_col(df, resolved.get('sku_id','wh_skuid'))
    qc = _find_query_col(list(df.columns))
    fc = _find_feedback_col(list(df.columns))
    print(f"   Cols: outlet={oc}, sku={sc}, queries={qc}, feedback={fc}")
    if not oc or not sc or not fc:
        print("   WARNING: Missing columns — suppression skipped")
        return {}
    suppression = {}
    n = 0
    for _, row in df.iterrows():
        fb = row.get(fc)
        if not _is_false_alarm_feedback(fb): continue
        oid = str(row.get(oc,'')).strip()
        sku = str(row.get(sc,'')).strip()
        if not oid or not sku or oid=='nan' or sku=='nan': continue
        base = oid + sku
        raw_q = str(row.get(qc,'')) if qc else ''
        mns = [normalize(m) for m in raw_q.split(',')
               if m.strip() and m.strip().lower()!='nan']
        if base not in suppression: suppression[base] = {}
        fb_str = str(fb).strip()
        for mn in mns:
            suppression[base][mn] = fb_str
        if not mns:
            suppression[base]['_any'] = fb_str
        n += 1
    print(f"   {n} False Alarm records ({len(suppression)} outlet-SKUs)")
    return suppression

def apply_suppression(final_df, suppression_set, resolved):
    if not suppression_set or final_df.empty: return final_df
    oc = resolved.get('outlet_id')
    sc = resolved.get('sku_id')
    prev_fb = []
    for _, row in final_df.iterrows():
        if oc and sc and oc in row.index and sc in row.index:
            bc = str(row[oc]).strip() + str(row[sc]).strip()
        else:
            bc = str(row.get('Comb','')).strip()
        if bc not in suppression_set:
            prev_fb.append(np.nan); continue
        comb_fb = suppression_set[bc]
        metrics = [(m.strip(), normalize(m))
                   for m in str(row.get('Queries','')).split(',') if m.strip()]
        texts = []
        for mr, mn in metrics:
            if mn in comb_fb:   texts.append(f"{mr}: {comb_fb[mn]}")
            elif '_any' in comb_fb: texts.append(f"{mr}: {comb_fb['_any']}")
        prev_fb.append(' | '.join(texts) if texts else np.nan)
    final_df = final_df.copy()
    final_df['Previous Month Feedback'] = prev_fb
    n = final_df['Previous Month Feedback'].notna().sum()
    print(f"   {'🔁' if n>0 else 'ℹ️ '} {n} record(s) have prior feedback")
    return final_df


# ═══════════════════════════════════════════════════════════════
# PART 5D: NEW BASELINE / SUSTAINED TREND (stock+purchase only)
# ═══════════════════════════════════════════════════════════════

def _is_new_baseline(curr_val, prev_val, hist_vals, metric_name):
    if pd.isna(curr_val) or pd.isna(prev_val) or curr_val==0 or prev_val==0: return False,''
    clean = [v for v in hist_vals if pd.notna(v) and v!=0]
    if len(clean) < 2: return False,''
    hm = float(np.mean(clean))
    if hm == 0: return False,''
    digits = len(str(int(abs(hm)))) if int(abs(hm))>0 else 1
    t_high,t_low = {1:(7,1/7),2:(5,1/5),3:(4,1/4)}.get(min(digits,3),(3,1/3))
    pr = prev_val/hm
    if not ((pr>t_high) or (pr<t_low)): return False,''
    cr = curr_val/prev_val
    if not (0.5<=cr<=2.0): return False,''
    direction = 'drop' if prev_val<hm else 'rise'
    reason = (f"{metric_name}: New trend — sustained {direction} "
              f"(avg was {hm:.0f}, became {prev_val:.0f} last month, now {curr_val:.0f})")
    return True, reason

def apply_new_baseline_check(final_df, df_list, all_months, resolved, stock_purchase_cols):
    if final_df.empty or not stock_purchase_cols: return final_df
    oc, sc = resolved.get('outlet_id'), resolved.get('sku_id')
    curr_m, prev_m = all_months[-1], all_months[-2]
    PF = 'Previous Month Feedback'
    final_df = final_df.copy()
    if PF not in final_df.columns: final_df[PF] = np.nan
    lookup = {}
    for mo, dm in zip(all_months, df_list):
        if oc not in dm.columns or sc not in dm.columns: continue
        for _, r in dm.iterrows():
            bc = str(r[oc]).strip()+str(r[sc]).strip()
            if bc not in lookup: lookup[bc] = {}
            for col in stock_purchase_cols:
                ac = resolve_col(dm, col)
                if ac and ac in r.index:
                    val = pd.to_numeric(r[ac], errors='coerce')
                    if col not in lookup[bc]: lookup[bc][col] = {}
                    lookup[bc][col][mo] = val
    count = 0
    for idx, row in final_df.iterrows():
        if pd.notna(row.get(PF)): continue
        qs = str(row.get('Queries',''))
        if oc and sc and oc in row.index and sc in row.index:
            bc = str(row[oc]).strip()+str(row[sc]).strip()
        else:
            bc = str(row.get('Comb','')).strip()
        if bc not in lookup: continue
        reasons = []
        for metric in [m.strip() for m in qs.split(',') if m.strip()]:
            if metric not in stock_purchase_cols:
                if not any(normalize(metric)==normalize(s) for s in stock_purchase_cols):
                    continue
            md = lookup[bc].get(metric, {})
            if not md:
                for s, d in lookup[bc].items():
                    if normalize(s)==normalize(metric): md=d; break
            curr_v = md.get(curr_m)
            prev_v = md.get(prev_m)
            hist_v = [md[m] for m in all_months[:-2] if m in md]
            ok, reason = _is_new_baseline(curr_v, prev_v, hist_v, metric)
            if ok: reasons.append(reason)
        if reasons:
            final_df.at[idx, PF] = ' | '.join(reasons)
            count += 1
    if count: print(f"   📈 {count} record(s) — new baseline/trend")
    return final_df


# ═══════════════════════════════════════════════════════════════
# PART 5E: SPIKE-AND-RETURN (all metrics)
# ═══════════════════════════════════════════════════════════════

def is_valid_metric_value(value, is_price=False, skip_val_1=True):
    value = pd.to_numeric(value, errors='coerce')
    if pd.isna(value) or value==0: return False
    if is_price and skip_val_1 and value==1: return False
    return True

def _clean_hist_range(hist_vals):
    clean = [v for v in hist_vals if pd.notna(v) and v!=0]
    if len(clean) < 4: return None, None, []
    med = np.median(clean)
    mad = np.median([abs(v-med) for v in clean])
    if mad == 0: return med*0.9, med*1.1, []
    lo = med - 3*mad
    hi = med + 3*mad
    removed = [v for v in clean
               if (v<lo or v>hi) and (v>med*2.0 or v<med*0.5)]
    if not removed: return None, None, []
    return lo, hi, removed

def _is_spike_and_return(curr_val, month_data, current_month, metric_name,
                          is_price=False):
    if not is_valid_metric_value(curr_val, is_price=is_price): return False,''
    hist = {m:v for m,v in month_data.items()
            if m!=current_month and is_valid_metric_value(v, is_price=is_price)}
    if len(hist) < 3: return False,''
    lo, hi, removed = _clean_hist_range(list(hist.values()))
    if lo is None or not removed: return False,''
    if not (lo<=curr_val<=hi): return False,''
    direction = 'spike' if any(v>hi for v in removed) else 'dip'
    norm_min = float(min(v for v in hist.values() if v not in removed)) if \
        [v for v in hist.values() if v not in removed] else lo
    norm_max = float(max(v for v in hist.values() if v not in removed)) if \
        [v for v in hist.values() if v not in removed] else hi
    spike_desc = ', '.join(
        f"{m} (value {v:.0f})" for m,v in
        [(m,v) for m,v in month_data.items() if v in removed][:2])
    reason = (f"{metric_name}: Spike-and-return — temporary {direction} "
              f"({spike_desc}). Normal range ~{norm_min:.0f}–{norm_max:.0f}, "
              f"current {curr_val:.0f} is normal.")
    return True, reason

def apply_spike_and_return_check(final_df, df_list, all_months, resolved,
                                  all_metric_cols, price_cols):
    if final_df.empty or not all_metric_cols: return final_df
    oc, sc = resolved.get('outlet_id'), resolved.get('sku_id')
    curr_m = all_months[-1]
    PF = 'Previous Month Feedback'
    final_df = final_df.copy()
    if PF not in final_df.columns: final_df[PF] = np.nan
    lookup = {}
    for mo, dm in zip(all_months, df_list):
        if oc not in dm.columns or sc not in dm.columns: continue
        for _, r in dm.iterrows():
            bc = str(r[oc]).strip()+str(r[sc]).strip()
            if bc not in lookup: lookup[bc] = {}
            for col in all_metric_cols:
                ac = resolve_col(dm, col)
                if ac and ac in r.index:
                    val = pd.to_numeric(r[ac], errors='coerce')
                    if col not in lookup[bc]: lookup[bc][col] = {}
                    lookup[bc][col][mo] = val
    price_set = set(price_cols)
    count = 0
    for idx, row in final_df.iterrows():
        if pd.notna(row.get(PF)): continue
        qs = str(row.get('Queries',''))
        if oc and sc and oc in row.index and sc in row.index:
            bc = str(row[oc]).strip()+str(row[sc]).strip()
        else:
            bc = str(row.get('Comb','')).strip()
        if bc not in lookup: continue
        reasons = []
        for metric in [m.strip() for m in qs.split(',') if m.strip()]:
            if metric in ('Negative sales','Negative Profit','No Profit',
                          'Facings exceeds unit stock'):
                continue
            md = lookup[bc].get(metric, {})
            if not md:
                for mc, d in lookup[bc].items():
                    if normalize(mc)==normalize(metric): md=d; break
            if not md: continue
            curr_v = md.get(curr_m)
            is_pc = metric in price_set or any(normalize(metric)==normalize(p) for p in price_set)
            if not is_valid_metric_value(curr_v, is_price=is_pc): continue
            ok, reason = _is_spike_and_return(curr_v, md, curr_m, metric, is_price=is_pc)
            if ok: reasons.append(reason)
        if reasons:
            final_df.at[idx, PF] = ' | '.join(reasons)
            count += 1
    if count: print(f"   🔄 {count} record(s) — spike-and-return")
    return final_df


# ═══════════════════════════════════════════════════════════════
# PART 6: IQR HELPERS
# ═══════════════════════════════════════════════════════════════

def clean_values(vals, is_price=False, skip_val_1=True):
    vals = pd.to_numeric(pd.Series(vals), errors='coerce')
    vals = vals[vals.notna() & (vals!=0)]
    if is_price and skip_val_1:
        vals = vals[vals!=1]
    return vals

def get_iqr_bounds(vals, is_price=False, is_stock_stage1=False, stage=1):
    multiplier = 5 if is_stock_stage1 else 1.5
    vals = clean_values(vals, is_price=is_price)
    if len(vals) < 2: return np.nan, np.nan
    q1,q3 = np.percentile(vals,[25,75])
    iqr = q3-q1
    return q1-multiplier*iqr, q3+multiplier*iqr

def digit_multiplier(val):
    if pd.isna(val) or val==0: return 1.5
    digits = len(str(int(abs(val)))) if int(abs(val))>0 else 1
    if digits in [2,3,4]: return 2.0
    if digits in [5,6]:   return 1.5
    return 1.5


# ═══════════════════════════════════════════════════════════════
# PART 7: COMB IQR DETECTION — VECTORIZED
# ═══════════════════════════════════════════════════════════════

def run_comb_detection(df_list, all_months, sheet_names, current_month, prev_month,
                       columns_to_check, resolved, config, comb_type='stock'):
    oid_col    = resolved['outlet_id']
    sku_col    = resolved['sku_id']
    comb_field = 'Comb_price' if comb_type=='price' else 'Comb_stock'
    is_price   = comb_type=='price'
    skip_1     = config.get('skip_price_val_1', True)
    p_low      = config.get('price_low', 0.25)
    p_high     = config.get('price_high', 2.0)

    all_codes = pd.concat([df[[comb_field, oid_col, sku_col]]
                           for df in df_list if comb_field in df.columns])
    all_codes = (all_codes.dropna(subset=[comb_field])
                          .drop_duplicates(comb_field)
                          .reset_index(drop=True))

    results          = {}
    non_flagged_data = {}

    for target_col in columns_to_check:
        actual_col = resolve_col(df_list[0], target_col)
        if actual_col is None:
            print(f"   WARNING: '{target_col}' not found — skipping")
            continue

        is_stk = actual_col == resolved.get('stock_col')
        monthly_data, outlet_map, sku_map = {}, {}, {}

        for sheet, df in zip(all_months, df_list):
            if comb_field not in df.columns or actual_col not in df.columns: continue
            df_nn = df.dropna(subset=[comb_field])
            monthly_data[sheet] = df_nn.set_index(comb_field)[actual_col]
            outlet_map.update(df_nn.set_index(comb_field)[oid_col].to_dict())
            sku_map.update(df_nn.set_index(comb_field)[sku_col].to_dict())

        combined = pd.DataFrame(index=all_codes[comb_field].drop_duplicates())
        for month in all_months:
            data = monthly_data.get(month)
            if data is not None:
                data = data[~data.index.duplicated(keep='first')]
                combined[month] = data.reindex(combined.index)

        combined.insert(0, sku_col,    combined.index.map(sku_map))
        combined.insert(0, oid_col,    combined.index.map(outlet_map))
        combined.insert(0, comb_field, combined.index)

        hist_mat = combined[sheet_names].to_numpy(dtype=float).copy()
        hist_mat[hist_mat == 0] = np.nan
        if is_price and skip_1:
            hist_mat[hist_mat == 1] = np.nan

        multiplier = 5.0 if is_stk else 1.5
        with np.errstate(all='ignore'):
            q1 = np.nanpercentile(hist_mat, 25, axis=1)
            q3 = np.nanpercentile(hist_mat, 75, axis=1)
            iqr = q3 - q1
            lo_all = q1 - multiplier * iqr
            hi_all = q3 + multiplier * iqr
            tol_min = lo_all - np.abs(lo_all * 0.5)
            tol_max = hi_all + np.abs(hi_all * 0.5)

        curr_v = combined[current_month].to_numpy(dtype=float)
        curr_v_clean = curr_v.copy()
        curr_v_clean[curr_v_clean == 0] = np.nan
        if is_price and skip_1:
            curr_v_clean[curr_v_clean == 1] = np.nan

        valid_curr = ~np.isnan(curr_v_clean)
        valid_hist = ~np.isnan(tol_min)

        s1 = np.where(
            valid_curr & valid_hist,
            np.where((curr_v_clean < tol_min) | (curr_v_clean > tol_max), 'True', 'False'),
            ''
        )
        combined['_s1'] = s1

        last3 = all_months[-4:-1]
        last3_available = [m for m in last3 if m in combined.columns]

        if len(last3_available) >= 2:
            l3_mat = combined[last3_available].to_numpy(dtype=float).copy()
            l3_mat[l3_mat == 0] = np.nan
            if is_price and skip_1:
                l3_mat[l3_mat == 1] = np.nan

            with np.errstate(all='ignore'):
                lq1 = np.nanpercentile(l3_mat, 25, axis=1)
                lq3 = np.nanpercentile(l3_mat, 75, axis=1)
                liqr = lq3 - lq1
                lo3  = lq1 - 1.5 * liqr
                hi3  = lq3 + 1.5 * liqr
        else:
            lo3 = hi3 = np.full(len(combined), np.nan)

        if is_price:
            t_lo_arr = np.full(len(combined), p_low)
            t_hi_arr = np.full(len(combined), p_high)
        else:
            all_mat = combined[all_months].to_numpy(dtype=float).copy()
            all_mat[all_mat == 0] = np.nan
            row_means = np.nanmean(all_mat, axis=1)
            def _thresh_vec(means):
                t_hi = np.full(len(means), 3.0)
                t_lo = np.full(len(means), 1/3)
                for digits, (th, tl) in [(1,(7,1/7)),(2,(5,1/5)),(3,(4,1/4))]:
                    lo_d = 10**(digits-1)
                    hi_d = 10**digits
                    mask = (means >= lo_d) & (means < hi_d)
                    t_hi[mask] = th
                    t_lo[mask] = tl
                return t_hi, t_lo
            t_hi_arr, t_lo_arr = _thresh_vec(row_means)

        prev_v = combined[prev_month].to_numpy(dtype=float)

        s1_flags = combined['_s1'].to_numpy()
        s2 = np.empty(len(combined), dtype=object)
        s2[:] = ''

        run_mask = np.isin(s1_flags, ['True', 'False'])
        valid_lower = ~np.isnan(lo3) & ~np.isnan(hi3)

        below = (curr_v_clean < lo3 * t_lo_arr)
        above = (curr_v_clean > hi3 * t_hi_arr)

        s2[run_mask & valid_curr & valid_lower] = np.where(
            (below | above)[run_mask & valid_curr & valid_lower],
            'True', 'False'
        )
        s2[~run_mask] = ''

        flag_col = f'{target_col}_Outlier (3 Months)'
        combined[flag_col] = s2

        if not is_price:
            target_norm = normalize(target_col)
            stock_norm = normalize(resolved.get('stock_col', ''))
            purchase_norm = normalize(resolved.get('purchase_col', ''))

            low_limit = None
            if target_norm == stock_norm:
                low_limit = LOW_STOCK_QUERY_LIMIT
            elif target_norm == purchase_norm:
                low_limit = LOW_PURCHASE_QUERY_LIMIT

            low_months = [m for m in all_months[-4:] if m in combined.columns]
            if low_limit is not None and len(low_months) == 4:
                low_mat = combined[low_months].to_numpy(dtype=float).copy()
                low_mat[low_mat == 0] = np.nan
                low_valid = ~np.isnan(low_mat)
                below_or_missing = (~low_valid) | (low_mat < low_limit)
                low_suppress = low_valid.any(axis=1) & below_or_missing.all(axis=1)
                combined.loc[low_suppress, flag_col] = 'False'

        if not is_price and len(all_months) >= 2:
            sp_mat = combined[all_months].to_numpy(dtype=float).copy()
            sp_mat[sp_mat == 0] = np.nan
            sp_valid = ~np.isnan(sp_mat)

            curr_idx = len(all_months) - 1
            prev_idx = len(all_months) - 2
            older_idx = list(range(0, max(0, len(all_months) - 2)))

            has_curr_prev = sp_valid[:, curr_idx] & sp_valid[:, prev_idx]
            no_older_data = ~sp_valid[:, older_idx].any(axis=1) if older_idx else np.ones(len(combined), dtype=bool)
            new_outlet_suppress = has_curr_prev & no_older_data
            combined.loc[new_outlet_suppress, flag_col] = 'False'

            # Stock/Purchase only: if the previous month is blank/zero,
            # suppress the already-detected outlier. Do not apply this to prices.
            prev_blank_or_zero = ~sp_valid[:, prev_idx]
            combined.loc[(combined[flag_col] == 'True') & prev_blank_or_zero, flag_col] = 'False'

        if is_price:
            def _valid_p(arr):
                a = arr.copy().astype(float)
                mask = ~np.isnan(a) & (a != 0)
                if skip_1: mask &= (a != 1)
                return mask

            prev_valid = _valid_p(prev_v)
            curr_valid = _valid_p(curr_v_clean)
            no_flag    = np.isin(combined[flag_col].to_numpy(), ['False',''])

            ov1_mask = no_flag & prev_valid & curr_valid
            mults    = np.where((np.floor(np.log10(np.abs(prev_v+1e-9))).astype(int)+1).clip(1,6)
                                  .reshape(-1) <= 4, 2.0, 1.5)
            ov1_lo   = np.where(prev_valid, prev_v / mults, np.nan)
            ov1_hi   = np.where(prev_valid, prev_v * mults, np.nan)
            oob      = ov1_mask & (curr_v_clean < ov1_lo) | (curr_v_clean > ov1_hi)
            combined.loc[oob, flag_col] = 'True'

            if len(all_months) >= 3:
                pprev_v = combined[all_months[-3]].to_numpy(dtype=float)
                jt_mask = oob
                lo_r = np.minimum(pprev_v * 0.5, pprev_v / 0.5)
                hi_r = np.maximum(pprev_v * 0.5, pprev_v / 0.5)
                inside = jt_mask & (curr_v_clean >= lo_r) & (curr_v_clean <= hi_r)
                combined.loc[inside, flag_col] = 'False'

        _cf  = comb_field
        _res = combined[[_cf, flag_col]].copy()
        if _res.index.name == _cf:
            _res = _res.reset_index(drop=True)
        results[target_col] = _res

        # NF audit sheet: include ALL records, not only non-outliers.
        # The final column clearly shows whether this metric was detected as an outlier.
        audit_df = combined.copy()
        nf_cols = [oid_col, sku_col, comb_field]
        month_value_cols = []
        for mo in all_months:
            if mo in combined.columns:
                month_col = f'{target_col} ({mo})'
                audit_df[month_col] = combined[mo]
                month_value_cols.append(month_col)

        audit_df['Is Outlier'] = audit_df[flag_col].eq('True')
        non_flagged_data[target_col] = audit_df[
            nf_cols + month_value_cols + ['Is Outlier']
        ].reset_index(drop=True)

    return results, non_flagged_data


# ═══════════════════════════════════════════════════════════════
# PART 8: COMB2 (SKU + CHANNEL) IQR DETECTION
# ═══════════════════════════════════════════════════════════════

def run_comb2_detection(df_list, all_months, current_month, price_cols, resolved, config):
    oid_col   = resolved['outlet_id']
    sku_col   = resolved['sku_id']
    chn_col   = resolved.get('channel')
    use_units = config.get('use_units_in_price_comb', False)
    unt_col   = resolved.get('units_col')
    if not chn_col:
        print("   WARNING: No channel column — Comb2 skipped")
        return []
    comb2_rows = []
    def _build_c2(df):
        base = df[sku_col].astype(str) + df[chn_col].astype(str)
        if use_units and unt_col and unt_col in df.columns:
            return base + df[unt_col].astype(str)
        return base
    for target_col in price_cols:
        ac = resolve_col(df_list[0], target_col)
        if ac is None: continue
        hist_parts = []
        for df in df_list[:-1]:
            if ac not in df.columns: continue
            tmp = df.copy()
            tmp[ac] = pd.to_numeric(tmp[ac], errors='coerce')
            tmp['Comb2'] = _build_c2(tmp)
            tmp['Comb_price'] = (tmp[oid_col].astype(str) + tmp[sku_col].astype(str)
                + (tmp[unt_col].astype(str) if use_units and unt_col and unt_col in tmp.columns else ''))
            hist_parts.append(tmp[['Comb2', ac]])
        if not hist_parts: continue
        hist_df = pd.concat(hist_parts).dropna()
        def _iqr_b(v):
            v = v[(v!=0)&(v!=1)].dropna()
            if len(v)<2: return pd.Series([np.nan,np.nan])
            q1,q3 = np.percentile(v,[25,75])
            return pd.Series([q1-1.5*(q3-q1), q3+1.5*(q3-q1)])
        limits = hist_df.groupby('Comb2')[ac].apply(_iqr_b).reset_index()
        limits.columns = ['Comb2', 'Lower', 'Upper']

        # keep only required columns to avoid duplicate merge columns
        limits = limits[['Comb2', 'Lower', 'Upper']]
        curr = df_list[-1].copy()
        curr[ac] = pd.to_numeric(curr[ac], errors='coerce')
        curr['Comb2'] = _build_c2(curr)
        curr['Comb_price'] = (curr[oid_col].astype(str) + curr[sku_col].astype(str)
            + (curr[unt_col].astype(str) if use_units and unt_col and unt_col in curr.columns else ''))
        curr = curr[(curr[ac]!=0)&curr[ac].notna()]
        if config.get('skip_price_val_1',True): curr = curr[curr[ac]!=1]
        joined = curr.merge(limits, on='Comb2', how='left', suffixes=('', '_lim'))
        def _is_out(r):
            v,lo,hi = r[ac],r['Lower'],r['Upper']
            if pd.isna(v) or pd.isna(lo) or pd.isna(hi): return False
            if lo==0 or hi==0: return False
            pct = 150
            if v<lo: return abs(v-lo)/abs(lo)*100>pct
            if v>hi: return abs(v-hi)/abs(hi)*100>pct
            return False
        outliers = joined[joined.apply(_is_out,axis=1)].copy()
        outliers = outliers[outliers[ac].notna()]
        if outliers.empty:
            continue
        skip_1_flag = config.get('skip_price_val_1', True)

        def _valid_price(v):
            v = pd.to_numeric(v, errors='coerce')
            if pd.isna(v) or v == 0:
                return None
            if skip_1_flag and v == 1:
                return None
            return v

        def _get_price_for_comb(df_x, comb_price):
            if 'Comb_price' not in df_x.columns or ac not in df_x.columns:
                return None
            m = df_x[df_x['Comb_price'] == comb_price]
            if m.empty:
                return None
            return _valid_price(m.iloc[0][ac])

        def _stable(r):
            cp = r['Comb_price']
            cv = _valid_price(r[ac])
            if cv is None:
                return False
            recent_valid = []
            for dx in reversed(df_list[:-1][-2:]):
                v = _get_price_for_comb(dx, cp)
                if v is not None:
                    recent_valid.append(v)
            if recent_valid:
                return any(v == cv for v in recent_valid)
            older_valid = []
            for dx in reversed(df_list[:-3]):
                v = _get_price_for_comb(dx, cp)
                if v is not None:
                    older_valid.append(v)
                if len(older_valid) == 2:
                    break
            if len(older_valid) < 2:
                return False
            return all(v == cv for v in older_valid)
        outliers = outliers[~outliers.apply(_stable, axis=1)]
        if outliers.empty:
            continue
        outliers['Current Outlier Value'] = outliers[ac]
        outliers['Flagged Column'] = target_col
        comb_to_row = {}
        for df in reversed(df_list):
            if 'Comb_price' not in df.columns: continue
            for _,r in df.iterrows():
                k = str(r['Comb_price'])
                if k not in comb_to_row: comb_to_row[k] = r.to_dict()
        for _,row in outliers.iterrows():
            cp = str(row['Comb_price'])
            fc = row['Flagged Column']
            fr = comb_to_row.get(cp)
            if fr is None: continue
            fr = dict(fr)
            fr['Comb'] = cp
            fr['Queries'] = fc
            fr['Combination'] = 'Comb2'
            fr[f'{fc} ({current_month})'] = row['Current Outlier Value']
            for mo, dm in zip(all_months, df_list):
                if 'Comb_price' not in dm.columns: continue
                mm = dm[dm['Comb_price']==cp]
                if not mm.empty and ac in mm.columns:
                    fr[f'{fc} ({mo})'] = mm.iloc[0][ac]
            comb2_rows.append(fr)
    print(f"   Comb2: {len(comb2_rows)} flagged")
    return comb2_rows


# ═══════════════════════════════════════════════════════════════
# PART 9: BUILD COMB OUTPUT ROWS
# ═══════════════════════════════════════════════════════════════

def build_comb_rows(comb_results, df_list, all_months, current_month,
                    resolved, columns_to_check, sales_dict, comb_type='stock'):
    comb_field = 'Comb_price' if comb_type=='price' else 'Comb_stock'
    oid_col    = resolved['outlet_id']
    sku_col    = resolved['sku_id']
    all_combs  = pd.concat([df[[comb_field]].drop_duplicates()
                             for df in df_list if comb_field in df.columns]
                           ).drop_duplicates(comb_field).reset_index(drop=True)
    merged = all_combs.copy()
    for col, df in comb_results.items():
        if df.index.name==comb_field or comb_field in getattr(df.index,'names',[]):
            df = df.reset_index()
        fcn  = f'{col}_Outlier (3 Months)'
        keep = [c for c in [comb_field, fcn] if c in df.columns]
        merged = pd.merge(merged, df[keep], on=comb_field, how='left')
    def _bq(row):
        return ', '.join([c for c in columns_to_check
                          if row.get(f'{c}_Outlier (3 Months)')=='True'])
    merged['Queries']     = merged.apply(_bq, axis=1)
    merged                = merged[merged['Queries']!='']
    merged['Combination'] = 'Comb'
    comb_to_row = {}
    for df in reversed(df_list):
        if comb_field not in df.columns: continue
        for _,row in df.iterrows():
            k = str(row[comb_field])
            if k not in comb_to_row: comb_to_row[k] = row.to_dict()
    final_rows = []
    for _, row in merged.iterrows():
        cp = row[comb_field]
        fr = comb_to_row.get(str(cp))
        if fr is None: continue
        fr = dict(fr)
        fr['Comb']        = str(cp)
        fr['Queries']     = row['Queries']
        fr['Combination'] = row['Combination']
        flagged = [c.strip() for c in row['Queries'].split(', ')]
        if 'Negative sales' in flagged:
            fr['Sales'] = sales_dict.get(str(cp), np.nan)
        for col in flagged:
            actual = resolve_col(df_list[-1], col)
            if not actual: continue
            for mo, dm in zip(all_months, df_list):
                if comb_field not in dm.columns: continue
                mm = dm[dm[comb_field].astype(str)==str(cp)]
                if not mm.empty and actual in mm.columns:
                    fr[f'{col} ({mo})'] = mm.iloc[0][actual]
        final_rows.append(fr)
    return final_rows


# ═══════════════════════════════════════════════════════════════
# PART 10: OUTPUT FORMATTING
# ═══════════════════════════════════════════════════════════════

def reorder_output(df, months, col_order):
    if df.empty: return df
    suf     = tuple(f'({m})' for m in months)
    base    = [c for c in df.columns if not c.endswith(suf) and c!='Sales']
    mc_all  = []
    for metric in col_order:
        for mo in months:
            cn = f'{metric} ({mo})'
            mc_all.append(cn)
            if cn not in df.columns: df[cn] = np.nan
    if 'Sales' not in df.columns: df['Sales'] = np.nan
    out_rows = []
    for _, row in df.iterrows():
        ro  = row[base].to_dict()
        qs  = str(row.get('Queries',''))
        ro['Sales'] = row.get('Sales',np.nan) if 'Negative sales' in qs else np.nan
        for metric in col_order:
            for mo in months:
                cn   = f'{metric} ({mo})'
                show = (metric in qs
                        or ('Negative sales' in qs and any(k in normalize(metric)
                            for k in ['stock','purchase','purchases']))
                        or any(k in qs for k in ['Negative Profit','No Profit']
                               if metric in ['Selling Price per Sku',
                                             'Buying Price per Sku','Selling Price']))
                ro[cn] = row.get(cn, np.nan) if show else np.nan
        out_rows.append(ro)
    return pd.DataFrame(out_rows)[base+['Sales']+mc_all]


# ═══════════════════════════════════════════════════════════════
# PART 10A: FINAL QUERIES SHEET BUILDER
# ═══════════════════════════════════════════════════════════════

def _fq_find(keywords, columns):
    """Return the first column whose normalized name matches any keyword. Silent if none found."""
    for kw in keywords:
        kn = normalize(kw)
        for col in columns:
            cn = normalize(col)
            if cn == kn or kn in cn:
                return col
    return None


def build_final_queries(df_act, all_months, resolved):
    """
    Re-structure the Active Queries rows into the Final Queries column layout:

      Batch (if present) | Out Number | Prod Code | Outlet Details |
      Product Description | Auditor |
      [Stock sub-cols as-is] | Total Stock |
      [Purchase sub-cols: P Day / PURCHW] | Purchase col (main) |
      [Price cols] | Units | Channel Type | Queries |
      [Prev-month history: prices first → stock → purchases]

    All column lookups are silent — missing columns are simply omitted.
    """
    if df_act.empty:
        return pd.DataFrame()

    df = df_act.copy()

    # Split base columns from "(Month)" history columns
    suf_tuples = tuple(f'({m})' for m in all_months)
    base_cols  = [c for c in df.columns if not c.endswith(suf_tuples)]
    month_cols = [c for c in df.columns if c.endswith(suf_tuples)]

    # Key resolved columns
    out_col      = resolved.get('outlet_id')
    sku_col      = resolved.get('sku_id')
    price_cols   = resolved.get('price_cols') or []
    buy_col      = resolved.get('buying_price_col')
    units_col    = resolved.get('units_col')
    stock_col    = resolved.get('stock_col')      # Total Stock
    purchase_col = resolved.get('purchase_col')   # Main purchase column

    # Columns that must never be pulled in as stock/purchase sub-cols
    META = {'Comb', 'Comb_stock', 'Comb_price', '_res_unit',
            'Queries', 'Combination', 'Previous Month Feedback', 'Sales'}
    if out_col:      META.add(out_col)
    if sku_col:      META.add(sku_col)
    if stock_col:    META.add(stock_col)
    if purchase_col: META.add(purchase_col)
    for pc in price_cols:
        META.add(pc)
    if buy_col:   META.add(buy_col)
    if units_col: META.add(units_col)

    # ── Batch (goes first) ──────────────────────────────────────
    batch_col = _fq_find(['batch', 'batchno', 'batch number', 'batch no'], base_cols)

    # ── Outlet Details ──────────────────────────────────────────
    # Use existing column if present, otherwise concatenate components with " - "
    od_col = _fq_find(['outletdetails', 'outlet details', 'outlet_details'], base_cols)
    if od_col:
        outlet_detail_series = df[od_col].astype(str).replace('nan', '')
    else:
        DETAIL_SEARCHES = [
            ['outletname',    'outlet name',    'outlet_name'],
            ['tradechannel',  'channeltype',    'channel type', 'trade channel'],
            ['greaterregion', 'greater region', 'region'],
            ['district'],
            ['town'],
        ]
        comp = []
        for searches in DETAIL_SEARCHES:
            c = _fq_find(searches, base_cols)
            if c and c not in META:
                comp.append(c)
        if out_col and out_col not in comp:
            comp.append(out_col)   # outlet number appended last

        def _concat(row):
            parts = [str(row[c]).strip() for c in comp
                     if pd.notna(row.get(c)) and str(row.get(c, 'nan')).strip() not in ('', 'nan')]
            return ' - '.join(parts)

        outlet_detail_series = df.apply(_concat, axis=1)

    # ── Product Description ─────────────────────────────────────
    desc_col = _fq_find(
        ['productdescription', 'skudescription', 'product description',
         'sku description', 'proddescription', 'prod description', 'description'],
        base_cols)

    # ── Auditor ─────────────────────────────────────────────────
    aud_col = _fq_find(
        ['auditorname', 'auditor', 'fwname', 'fw name', 'fieldworkername', 'fieldworker'],
        base_cols)

    # ── Stock sub-columns (front, back, warm, cold, cooler, facings…) ──
    # Include columns with stock-like keywords; exclude anything that
    # looks like a purchase, price, batch, or identity column.
    STOCK_KW   = {'warm', 'cold', 'cooler', 'bank', 'facing', 'facings',
                  'front', 'back', 'stock', 'unit', 'units'}
    STOCK_EXCL = {'purchase', 'pday', 'purchw', 'price', 'buying', 'selling',
                  'batch', 'outlet', 'sku', 'prod', 'audit', 'comb',
                  'description', 'channel', 'region', 'district', 'town',
                  'name', 'fw', 'feedback', 'sales', 'date', 'code',
                  'id', 'query', 'queries', 'total'}

    def _is_stock_sub(col):
        if col in META: return False
        n = normalize(col)
        if any(ex in n for ex in STOCK_EXCL): return False
        return any(kw in n for kw in STOCK_KW)

    stock_sub_cols = [c for c in base_cols if _is_stock_sub(c)]

    # ── Purchase sub-columns (P Day 1-N, PURCHW 1-N, etc.) ─────
    def _is_purchase_sub(col):
        if col in META: return False
        n = normalize(col)
        return any(kw in n for kw in ['pday', 'purchw', 'purchase', 'purchases'])

    purchase_sub_cols = [c for c in base_cols if _is_purchase_sub(c)]

    # ── Channel Type (for price section) ────────────────────────
    chan_type_col = _fq_find(
        ['channeltype', 'channel type', 'tradechannel', 'trade channel', 'channel'],
        base_cols)

    # ── Previous-month history columns — grouped by metric type ─
    all_price_names = (list(price_cols)
                       + ([buy_col] if buy_col else [])
                       + ([units_col] if units_col else []))

    def _metric_match(col_name, metric):
        return bool(metric) and normalize(metric) in normalize(col_name)

    prev_price_cols    = [c for c in month_cols if any(_metric_match(c, p) for p in all_price_names)]
    prev_stock_cols    = [c for c in month_cols if _metric_match(c, stock_col)]
    prev_purchase_cols = [c for c in month_cols if _metric_match(c, purchase_col)]

    # ── Assemble result dataframe ────────────────────────────────
    out = pd.DataFrame(index=df.index)

    def _take(src_col, label=None):
        """Copy src_col into out (optionally renamed). Silent if col is missing."""
        if src_col and src_col in df.columns:
            out[label or src_col] = df[src_col].values

    # 1. Batch (first column)
    _take(batch_col, 'Batch')

    # 2. Out Number
    _take(out_col, 'Out Number')

    # 3. Prod Code
    _take(sku_col, 'Prod Code')

    # 4. Outlet Details
    out['Outlet Details'] = outlet_detail_series.values

    # 5. Product Description
    _take(desc_col, 'Product Description')

    # 6. Auditor
    _take(aud_col, 'Auditor')

    # 7. Stock sub-columns (kept with their original names)
    for c in stock_sub_cols:
        _take(c)

    # 8. Total Stock
    _take(stock_col, 'Total Stock')

    # 9. Purchase sub-columns (P Day / PURCHW, kept with original names)
    for c in purchase_sub_cols:
        _take(c)

    # 10. Main purchase column
    _take(purchase_col)

    # 11. Price columns + units (original names)
    for c in price_cols:
        _take(c)
    _take(buy_col)
    _take(units_col)

    # 12. Channel Type
    _take(chan_type_col, 'Channel Type')

    # 13. Queries
    if 'Queries' in df.columns:
        out['Queries'] = df['Queries'].values

    # 14. Previous-month history: prices → stock → purchases
    for c in prev_price_cols + prev_stock_cols + prev_purchase_cols:
        _take(c)

    return out.reset_index(drop=True)


def _write_sheet(wb, ws, df, grey_rows=False, all_grey=False, freeze=True):
    PF = 'Previous Month Feedback'
    hdr = wb.add_format({'bold':True,'bg_color':'#4472C4','font_color':'white','border':1})
    grey = wb.add_format({'bg_color':'#D9D9D9','font_color':'#808080','italic':True})
    norm = wb.add_format({'bg_color':'#FFFFFF'})
    for ci, cn in enumerate(df.columns):
        ws.write(0, ci, cn, hdr)
    has_pf = PF in df.columns
    for ri, (_, row) in enumerate(df.iterrows(), start=1):
        if all_grey:
            fmt = grey
        elif grey_rows and has_pf:
            fmt = grey if (pd.notna(row.get(PF)) and str(row.get(PF)).strip()!='') else norm
        else:
            fmt = norm
        for ci, val in enumerate(row):
            ws.write(ri, ci, '' if pd.isna(val) else val, fmt)
    ws.set_column(0, len(df.columns)-1, 18)
    for wc in ['Queries','Previous Month Feedback']:
        if wc in df.columns:
            ws.set_column(df.columns.get_loc(wc), df.columns.get_loc(wc), 45)
    if freeze: ws.freeze_panes(1, 0)


def _export_with_formatting(df, output_file, non_flagged_dict=None,
                             df_list=None, all_months=None, resolved=None):
    """
    Export to Excel:
    Sheet 1 — All Queries        (every flagged record)
    Sheet 2 — Ignored Queries    (suppressed, grey)
    Sheet 3 — Active Queries     (field action required)
    Sheet 4 — Final Queries      (restructured Active Queries for field use)
    Sheet 5+ — Non-Flagged: X    (one per metric, all months, for audit)
    """
    PF = 'Previous Month Feedback'
    has_pf = PF in df.columns
    if has_pf:
        df_ign = df[df[PF].notna() & (df[PF].astype(str).str.strip()!='')].copy()
        df_act = df[df[PF].isna()  | (df[PF].astype(str).str.strip()=='')].copy()
    else:
        df_ign = pd.DataFrame(columns=df.columns)
        df_act = df.copy()

    # Build Final Queries from Active Queries rows
    fq_df = pd.DataFrame()
    if df_list is not None and all_months is not None and resolved is not None:
        try:
            fq_df = build_final_queries(df_act, all_months, resolved)
        except Exception as e:
            print(f"   WARNING: Could not build Final Queries sheet: {e}")

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        wb = writer.book

        # ── Sheet 1: All Queries ──────────────────────────────────
        df.to_excel(writer, sheet_name='All Queries', index=False)
        _write_sheet(wb, writer.sheets['All Queries'], df, grey_rows=True)

        # ── Sheet 2: Ignored Queries ──────────────────────────────
        if not df_ign.empty:
            df_ign.to_excel(writer, sheet_name='Ignored Queries', index=False)
            _write_sheet(wb, writer.sheets['Ignored Queries'], df_ign, all_grey=True)
        else:
            ws = wb.add_worksheet('Ignored Queries')
            ws.write(0, 0, 'No ignored queries.',
                     wb.add_format({'italic':True,'font_color':'#808080'}))

        # ── Sheet 3: Active Queries ───────────────────────────────
        df_act.to_excel(writer, sheet_name='Active Queries', index=False)
        _write_sheet(wb, writer.sheets['Active Queries'], df_act, grey_rows=False)

        # ── Sheet 4: Final Queries ────────────────────────────────
        if not fq_df.empty:
            fq_df.to_excel(writer, sheet_name='Final Queries', index=False)
            _write_sheet(wb, writer.sheets['Final Queries'], fq_df, grey_rows=False)
        else:
            ws = wb.add_worksheet('Final Queries')
            ws.write(0, 0, 'No active queries to display.',
                     wb.add_format({'italic':True,'font_color':'#808080'}))

        # ── Sheets 5+: Non-Flagged audit (one per metric column) ──
        if non_flagged_dict:
            for metric_col, nf_df in non_flagged_dict.items():
                if nf_df is None or nf_df.empty:
                    continue
                sheet_name = f"NF: {metric_col}"[:31]
                for ch in ['/', '\\', '?', '*', '[', ']', ':']:
                    sheet_name = sheet_name.replace(ch, ' ')
                try:
                    nf_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    _write_sheet(wb, writer.sheets[sheet_name], nf_df, grey_rows=False)
                except Exception as e:
                    print(f"   WARNING: Could not write sheet '{sheet_name}': {e}")

    print(f'   Sheet "All Queries":     {len(df):,} rows')
    print(f'   Sheet "Ignored Queries": {len(df_ign):,} rows')
    print(f'   Sheet "Active Queries":  {len(df_act):,} rows')
    print(f'   Sheet "Final Queries":   {len(fq_df):,} rows')
    if non_flagged_dict:
        for mc, nf in non_flagged_dict.items():
            if nf is not None and not nf.empty:
                print(f'   Sheet "NF: {mc[:20]}": {len(nf):,} rows')




# ═══════════════════════════════════════════════════════════════
# PART 10B: EXTRA QUERY MODULES — OUTLETS, SKU COUNTS, NEW/LOST
# ═══════════════════════════════════════════════════════════════

def _resolve_optional_col(df_list, candidates):
    """Find first matching optional column across sheets using exact/normalized names."""
    for df in df_list:
        for cand in candidates:
            found = resolve_col(df, cand)
            if found:
                return found
    return None


def _first_non_null_lookup(df_list, key_col, value_cols):
    """Build lookup by key_col using the first non-null value seen across all sheets."""
    value_cols = [c for c in value_cols if c]
    if not key_col or not value_cols:
        return None
    frames = []
    for df in df_list:
        existing = [c for c in [key_col] + value_cols if c in df.columns]
        if key_col in existing and len(existing) >= 2:
            frames.append(df[existing].copy())
    if not frames:
        return None
    all_df = pd.concat(frames, ignore_index=True)
    return (all_df.groupby(key_col, as_index=False)[value_cols]
                  .agg(lambda s: s.dropna().iloc[0] if len(s.dropna()) else np.nan))


def build_sku_counts(df_list, all_months, resolved):
    """Generic version of SKU COUNT code for every configured project."""
    outlet_col = resolved.get('outlet_id')
    sku_col = resolved.get('sku_id')
    stock_col = resolved.get('stock_col')
    purchase_col = resolved.get('purchase_col')
    required = [outlet_col, sku_col, stock_col, purchase_col]
    if not all(required):
        return pd.DataFrame()

    result = None
    for month, df in zip(all_months, df_list):
        if not set(required).issubset(df.columns):
            continue
        tmp = df.copy()
        tmp[stock_col] = pd.to_numeric(tmp[stock_col], errors='coerce').fillna(0)
        tmp[purchase_col] = pd.to_numeric(tmp[purchase_col], errors='coerce').fillna(0)
        filtered = tmp[(tmp[stock_col] > 0) | (tmp[purchase_col] > 0)]
        counts = filtered.groupby(outlet_col)[sku_col].nunique().reset_index()
        counts.rename(columns={sku_col: f'SKU COUNT_{month}'}, inplace=True)
        result = counts if result is None else result.merge(counts, on=outlet_col, how='outer')

    if result is None:
        return pd.DataFrame()

    # Keep only outlets active in the current sheet, same concept as the original SKU count code.
    df_last = df_list[-1].copy()
    if set(required).issubset(df_last.columns):
        df_last[stock_col] = pd.to_numeric(df_last[stock_col], errors='coerce').fillna(0)
        df_last[purchase_col] = pd.to_numeric(df_last[purchase_col], errors='coerce').fillna(0)
        valid_outlets = df_last[(df_last[stock_col] > 0) | (df_last[purchase_col] > 0)][outlet_col].unique()
        result = result[result[outlet_col].isin(valid_outlets)]

    month_cols_ordered = [f'SKU COUNT_{m}' for m in all_months if f'SKU COUNT_{m}' in result.columns]
    if not month_cols_ordered:
        return result
    last_month_col = month_cols_ordered[-1]
    previous_months = month_cols_ordered[-4:-1]

    def digits_number(x):
        try:
            if x == 0 or np.isnan(x):
                return 0
            return len(str(int(abs(x))))
        except Exception:
            return 0

    def check_outlier(avg, curr):
        digits = digits_number(avg)
        if np.isnan(avg) or pd.isna(curr):
            return False
        avg_rounded = round(avg)
        if digits == 1:
            threshold_low, threshold_high = 1 / 3, 3
        elif digits == 2:
            threshold_low, threshold_high = 1 / 2, 2
        elif digits == 3:
            threshold_low, threshold_high = 1 / 1.5, 1.5
        else:
            threshold_low, threshold_high = 1 / 1.2, 1.2
        return (curr < avg_rounded * threshold_low) or (curr > avg_rounded * threshold_high)

    def avg_without_zeros(row):
        vals = [row[m] for m in previous_months if m in row.index and pd.notna(row[m]) and row[m] != 0]
        return np.nan if len(vals) == 0 else np.mean(vals)

    result['Average_prev_3m'] = result.apply(avg_without_zeros, axis=1)
    result['Outliers'] = result.apply(lambda x: check_outlier(x['Average_prev_3m'], x[last_month_col]), axis=1)
    cols = [c for c in result.columns if c not in ['Average_prev_3m', 'Outliers']]
    return result[cols + ['Average_prev_3m', 'Outliers']]


def build_new_lost_products(df_list, all_months, resolved):
    """Generic version of New/Lost Products output."""
    outlet_col = resolved.get('outlet_id')
    sku_col = resolved.get('sku_id')
    stock_col = resolved.get('stock_col')
    purchase_col = resolved.get('purchase_col')
    required = [outlet_col, sku_col, purchase_col, stock_col]
    if not all(required):
        return pd.DataFrame()

    combined_rows = []
    outlets_present_by_month = {}
    for month, df in zip(all_months, df_list):
        if not set(required).issubset(df.columns):
            raise ValueError(f"Sheet {month} is missing required columns for New/Lost Products.")
        outlets_present_by_month[month] = set(df[outlet_col].dropna().astype(str))
        tmp = df[[outlet_col, sku_col, purchase_col, stock_col]].copy()
        tmp[outlet_col] = tmp[outlet_col].astype(str)
        tmp[sku_col] = tmp[sku_col].astype(str)
        tmp['COMB'] = tmp[outlet_col] + ' ' + tmp[sku_col]
        tmp['Month'] = month
        combined_rows.append(tmp[['COMB', outlet_col, sku_col, purchase_col, stock_col, 'Month']])

    if not combined_rows:
        return pd.DataFrame()
    all_data = pd.concat(combined_rows, ignore_index=True)

    def determine_value(row):
        purchase = pd.to_numeric(row[purchase_col], errors='coerce')
        stock = pd.to_numeric(row[stock_col], errors='coerce')
        if pd.isna(purchase) or pd.isna(stock):
            if (pd.notna(purchase) and purchase > 0) or (pd.notna(stock) and stock > 0):
                return 'Yes'
            return 'No'
        if (purchase > 0) or (stock > 0):
            return 'Yes'
        if (purchase == 0) and (stock == 0):
            return 'No'
        return 'No'

    all_data['Flag'] = all_data.apply(determine_value, axis=1)
    pivot_df = all_data.pivot_table(
        index=['COMB', outlet_col, sku_col],
        columns='Month',
        values='Flag',
        aggfunc='first'
    ).reset_index()
    pivot_df = pivot_df.reindex(columns=['COMB', outlet_col, sku_col] + all_months)
    pivot_df[all_months] = pivot_df[all_months].fillna('No')
    pivot_df[outlet_col] = pivot_df[outlet_col].astype(str)
    for month in all_months:
        present = outlets_present_by_month.get(month, set())
        pivot_df.loc[~pivot_df[outlet_col].isin(present), month] = ' '
    return pivot_df


def _outlet_thresholds(mean_val):
    int_part = int(abs(mean_val)) if pd.notna(mean_val) else 0
    digits = len(str(int_part)) if int_part > 0 else 1
    if digits == 1:
        return 1 / 7, 7
    if digits == 2:
        return 1 / 5, 5
    if digits == 3:
        return 1 / 4, 4
    if digits >= 4:
        return 1 / 3, 3
    return 1 / 2, 2


def _outlet_near(a, b, low=0.5, high=2.0):
    a = pd.to_numeric(a, errors='coerce')
    b = pd.to_numeric(b, errors='coerce')
    if pd.isna(a) or pd.isna(b) or a == 0 or b == 0:
        return False
    ratio = a / b
    return low <= ratio <= high


def _outlet_low_current_and_last3(row, metric_col, all_months, limit=50):
    check_months = all_months[-3:]
    if len(check_months) < 3:
        return False
    vals = []
    for month in check_months:
        val = pd.to_numeric(row.get(f'{metric_col} Sum ({month})', np.nan), errors='coerce')
        if pd.isna(val):
            val = 0
        vals.append(val)
    return all(v <= limit for v in vals)


def _outlet_current_near_last_month(row, metric_col, all_months):
    if len(all_months) < 2:
        return False
    current = row.get(f'{metric_col} Sum ({all_months[-1]})', np.nan)
    previous = row.get(f'{metric_col} Sum ({all_months[-2]})', np.nan)
    return _outlet_near(current, previous)


def _outlet_spike_and_return(row, metric_col, all_months):
    if len(all_months) < 4:
        return False

    current = pd.to_numeric(row.get(f'{metric_col} Sum ({all_months[-1]})', np.nan), errors='coerce')
    if pd.isna(current) or current == 0:
        return False

    # Test whether the last 1 or last 2 previous months were temporary spike/dip months,
    # and the current month has returned near the older baseline.
    for recent_count in (1, 2):
        if len(all_months) < recent_count + 3:
            continue

        recent_months = all_months[-(recent_count + 1):-1]
        baseline_months = all_months[:-(recent_count + 1)]

        baseline_vals = []
        for month in baseline_months:
            val = pd.to_numeric(row.get(f'{metric_col} Sum ({month})', np.nan), errors='coerce')
            if pd.notna(val) and val != 0:
                baseline_vals.append(float(val))

        recent_vals = []
        for month in recent_months:
            val = pd.to_numeric(row.get(f'{metric_col} Sum ({month})', np.nan), errors='coerce')
            if pd.notna(val) and val != 0:
                recent_vals.append(float(val))

        if len(baseline_vals) < 2 or len(recent_vals) != recent_count:
            continue

        baseline = float(np.median(baseline_vals))
        if baseline == 0:
            continue

        recent_high_spike = all(v > baseline * 2.0 for v in recent_vals)
        recent_low_dip = all(v < baseline * 0.5 for v in recent_vals)

        if (recent_high_spike or recent_low_dip) and _outlet_near(current, baseline):
            return True

    return False


def build_outlet_query_sheets(df_list, all_months, resolved):
    """Generic version of Outlet Stock / Outlet Purchases code."""
    outlet_col = resolved.get('outlet_id')
    sku_col = resolved.get('sku_id')
    stock_col = resolved.get('stock_col')
    purchase_col = resolved.get('purchase_col')
    if not all([outlet_col, sku_col, stock_col, purchase_col]):
        return {'Outlet Stock': pd.DataFrame(), 'Outlet Purchases': pd.DataFrame()}

    metrics = [(stock_col, 'Outlet Stock'), (purchase_col, 'Outlet Purchases')]
    metric_dfs = []

    for metric_col, _sheet_label in metrics:
        outlet_sums = {}
        for month, df in zip(all_months, df_list):
            if outlet_col not in df.columns or metric_col not in df.columns:
                continue
            tmp = df[[outlet_col, metric_col]].copy()
            tmp[metric_col] = pd.to_numeric(tmp[metric_col], errors='coerce').fillna(0)
            outlet_sums[month] = tmp.groupby(outlet_col)[metric_col].sum()
        if not outlet_sums:
            continue
        sums_df = pd.DataFrame(outlet_sums).reset_index()

        def flag_outlier(row):
            prev_vals = row[all_months[:-1]].values
            prev_vals = np.array([v for v in prev_vals if pd.notna(v) and v != 0])
            if len(prev_vals) < 2:
                return False, np.nan
            mean_val = np.mean(prev_vals)
            current = row[all_months[-1]]
            last_3 = all_months[-3:]
            last_3_vals = row[last_3].values
            if any(pd.isna(v) or v == 0 for v in last_3_vals):
                return False, mean_val
            if pd.isna(current):
                return False, mean_val
            low, high = _outlet_thresholds(mean_val)
            return (current < low * mean_val) or (current > high * mean_val), mean_val

        results = sums_df.apply(flag_outlier, axis=1, result_type='expand')
        sums_df[f'{metric_col}_Outlier'] = results[0]
        sums_df[f'{metric_col}_Prev_Mean'] = results[1]
        sums_df = sums_df.rename(columns={m: f'{metric_col} Sum ({m})' for m in all_months})

        def get_last2_mean(row):
            if len(all_months) < 3:
                return np.nan
            vals = []
            for month in [all_months[-2], all_months[-3]]:
                val = row.get(f'{metric_col} Sum ({month})', np.nan)
                val = pd.to_numeric(val, errors='coerce')
                if pd.notna(val) and val != 0:
                    vals.append(val)
            return np.nan if not vals else (vals[0] if len(vals) == 1 else np.mean(vals))

        sums_df[f'{metric_col}_Last2_Mean'] = sums_df.apply(get_last2_mean, axis=1)

        def flag_outlier_last2(row):
            if not row[f'{metric_col}_Outlier']:
                return False
            mean_last2 = row[f'{metric_col}_Last2_Mean']
            current = row.get(f'{metric_col} Sum ({all_months[-1]})', np.nan)
            if pd.isna(current) or pd.isna(mean_last2):
                return False
            low, high = _outlet_thresholds(mean_last2)
            return (current < low * mean_last2) or (current > high * mean_last2)

        sums_df[f'{metric_col}_Outlier_Last2'] = sums_df.apply(flag_outlier_last2, axis=1)

        # Outlet Stock/Purchases only: apply these extra suppressions only after
        # the outlet has already been detected as a query.
        already_query = sums_df[f'{metric_col}_Outlier_Last2'] == True
        low_current_last3 = sums_df.apply(
            lambda row: _outlet_low_current_and_last3(row, metric_col, all_months, limit=50),
            axis=1
        )
        spike_return = sums_df.apply(
            lambda row: _outlet_spike_and_return(row, metric_col, all_months),
            axis=1
        )
        current_near_last = sums_df.apply(
            lambda row: _outlet_current_near_last_month(row, metric_col, all_months),
            axis=1
        )
        sums_df.loc[already_query & (low_current_last3 | spike_return | current_near_last),
                    f'{metric_col}_Outlier_Last2'] = False

        metric_dfs.append((metric_col, _sheet_label, sums_df.set_index(outlet_col)))

    if not metric_dfs:
        return {'Outlet Stock': pd.DataFrame(), 'Outlet Purchases': pd.DataFrame()}

    merged = pd.concat([x[2] for x in metric_dfs], axis=1)
    merged = merged.loc[:, ~merged.columns.duplicated()].reset_index()

    # Optional columns, automatically resolved across projects.
    outlet_candidates = {
        'Outlet Details': ['OUTLET DETAILS', 'Outlet Details', 'Outlet Name', 'OutletName', 'Outlet'],
        'Channel Type': ['Channel Type', 'Trade Channel', 'TradeChannel', 'Channel'],
        'Auditor': ['AUDITOR', 'Auditor', 'Auditor Name', 'FW Name', 'Fieldworker'],
        'Batch': ['Batch1', 'Batch', 'batchid', 'Batch No', 'Batch Number'],
    }
    outlet_cols_present = []
    for _label, candidates in outlet_candidates.items():
        c = _resolve_optional_col(df_list, candidates)
        if c and c not in outlet_cols_present:
            outlet_cols_present.append(c)

    sku_desc_col = _resolve_optional_col(
        df_list,
        ['Product Description', 'Sku Description', 'SKU Description', 'DESCRIPTION', 'Prod Description']
    )
    sku_cols_present = [sku_desc_col] if sku_desc_col else []

    outlet_lookup = _first_non_null_lookup(df_list, outlet_col, outlet_cols_present)
    sku_lookup = _first_non_null_lookup(df_list, sku_col, sku_cols_present)

    def get_metric_month_cols(df, metric_col):
        cols = [c for c in df.columns if c.startswith(f'{metric_col} Sum (')]
        return sorted(cols, key=lambda x: all_months.index(x.split('(')[-1].rstrip(')')))

    def build_sku_all_months(metric_col):
        sku_wide = None
        for month, df in zip(all_months, df_list):
            if not set([outlet_col, sku_col, metric_col]).issubset(df.columns):
                continue
            dfm = df[[outlet_col, sku_col, metric_col]].copy()
            dfm[metric_col] = pd.to_numeric(dfm[metric_col], errors='coerce').fillna(0)
            dfm = (dfm.groupby([outlet_col, sku_col], as_index=False)[metric_col]
                     .sum()
                     .rename(columns={metric_col: f'{metric_col} ({month})'}))
            sku_wide = dfm if sku_wide is None else sku_wide.merge(dfm, on=[outlet_col, sku_col], how='outer')
        return sku_wide if sku_wide is not None else pd.DataFrame(columns=[outlet_col, sku_col])

    def build_outlet_product_sales_columns():
        if len(df_list) < 2:
            return pd.DataFrame(columns=[outlet_col, sku_col])

        curr_df = df_list[-1]
        prev_df = df_list[-2]

        required_curr = {outlet_col, sku_col, stock_col, purchase_col}
        required_prev = {outlet_col, sku_col, stock_col}
        if not required_curr.issubset(curr_df.columns) or not required_prev.issubset(prev_df.columns):
            return pd.DataFrame(columns=[outlet_col, sku_col])

        curr = curr_df[[outlet_col, sku_col, stock_col, purchase_col]].copy()
        prev = prev_df[[outlet_col, sku_col, stock_col]].copy()

        # Use the same matching key concept as Data Queries: Outlet ID + SKU ID.
        # Keep outlet/sku columns untouched for output/merging, but match through COMB.
        curr['_outlet_key'] = curr[outlet_col].astype(str).str.strip()
        curr['_sku_key'] = curr[sku_col].astype(str).str.strip()
        curr['_Comb_stock_key'] = curr['_outlet_key'] + curr['_sku_key']

        prev['_outlet_key'] = prev[outlet_col].astype(str).str.strip()
        prev['_sku_key'] = prev[sku_col].astype(str).str.strip()
        prev['_Comb_stock_key'] = prev['_outlet_key'] + prev['_sku_key']

        curr[stock_col] = pd.to_numeric(curr[stock_col], errors='coerce').fillna(0)
        curr[purchase_col] = pd.to_numeric(curr[purchase_col], errors='coerce').fillna(0)
        prev[stock_col] = pd.to_numeric(prev[stock_col], errors='coerce')

        curr = (curr.groupby(['_Comb_stock_key'], as_index=False)
                    .agg({
                        outlet_col: 'first',
                        sku_col: 'first',
                        stock_col: 'sum',
                        purchase_col: 'sum'
                    })
                    .rename(columns={
                        stock_col: 'Current Month Total Stock',
                        purchase_col: 'Current Month Purchases'
                    }))

        # Preserve missing previous month combinations as blank, not 0.
        prev = (prev.dropna(subset=['_Comb_stock_key'])
                    .groupby(['_Comb_stock_key'], as_index=False)
                    .agg({stock_col: 'sum'})
                    .rename(columns={stock_col: 'Previous Month Stock'}))

        sales_df = curr.merge(prev, on='_Comb_stock_key', how='left')

        # Outlet Stock/Purchases only:
        # If current stock or current purchases are blank, treat them as 0 and still calculate Sales.
        # If previous month stock is missing because the Outlet+SKU did not exist last month,
        # keep Previous Month Stock and Sales blank.
        sales_df['Current Month Purchases'] = pd.to_numeric(
            sales_df['Current Month Purchases'], errors='coerce'
        ).fillna(0)
        sales_df['Current Month Total Stock'] = pd.to_numeric(
            sales_df['Current Month Total Stock'], errors='coerce'
        ).fillna(0)
        sales_df['Previous Month Stock'] = pd.to_numeric(
            sales_df['Previous Month Stock'], errors='coerce'
        )

        sales_df['Sales'] = np.where(
            sales_df['Previous Month Stock'].notna(),
            sales_df['Previous Month Stock']
            + sales_df['Current Month Purchases']
            - sales_df['Current Month Total Stock'],
            np.nan
        )

        return sales_df[[outlet_col, sku_col,
                         'Previous Month Stock',
                         'Current Month Total Stock',
                         'Current Month Purchases',
                         'Sales']]

    outlet_product_sales = build_outlet_product_sales_columns()

    output = {}
    for metric_col, sheet_label in [(stock_col, 'Outlet Stock'), (purchase_col, 'Outlet Purchases')]:
        flag_col = f'{metric_col}_Outlier_Last2'
        if flag_col not in merged.columns:
            output[sheet_label] = pd.DataFrame()
            continue
        outlets_true = merged.loc[merged[flag_col] == True, [outlet_col]].drop_duplicates()
        month_sum_cols = get_metric_month_cols(merged, metric_col)
        outlet_level = merged.loc[merged[flag_col] == True, [outlet_col] + month_sum_cols].copy()
        if outlet_lookup is not None and outlet_cols_present:
            outlet_level = outlet_level.merge(outlet_lookup, on=outlet_col, how='left')
        sku_month_cols = [f'{metric_col} ({m})' for m in all_months]
        if outlet_level.empty:
            output[sheet_label] = pd.DataFrame(columns=[outlet_col] + outlet_cols_present + month_sum_cols + [sku_col] + sku_cols_present + sku_month_cols)
            continue
        sku_wide_all = build_sku_all_months(metric_col)
        sku_wide_all = sku_wide_all.merge(outlets_true, on=outlet_col, how='inner')
        if sku_lookup is not None and sku_cols_present:
            sku_wide_all = sku_wide_all.merge(sku_lookup, on=sku_col, how='left')
        for col in sku_month_cols:
            if col not in sku_wide_all.columns:
                sku_wide_all[col] = np.nan
        sku_wide_all = sku_wide_all[[outlet_col, sku_col] + sku_cols_present + sku_month_cols]

        extra_product_cols = []
        if outlet_product_sales is not None and not outlet_product_sales.empty:
            sku_wide_all = sku_wide_all.merge(outlet_product_sales, on=[outlet_col, sku_col], how='left')
            if sheet_label == 'Outlet Stock':
                extra_product_cols = ['Current Month Purchases', 'Sales']
            elif sheet_label == 'Outlet Purchases':
                extra_product_cols = ['Previous Month Stock', 'Current Month Total Stock', 'Sales']
            extra_product_cols = [c for c in extra_product_cols if c in sku_wide_all.columns]

        final_sheet = outlet_level.merge(sku_wide_all, on=outlet_col, how='inner')
        output[sheet_label] = final_sheet[[outlet_col] + outlet_cols_present + month_sum_cols + [sku_col] + sku_cols_present + extra_product_cols + sku_month_cols]
    return output


def _project_uses_outlet_query_ignore(project_name):
    """Only Nigeria MRA and Nigeria Ville should move duplicated stock/purchase rows to Ignored."""
    pn = normalize(project_name)
    allowed = {
        'ngmra', 'nigmra', 'nigeriamra',
        'nigeriavile', 'nigeriaville', 'ngvile', 'ngville'
    }
    return pn in allowed


def _apply_outlet_query_ignore(final_df, outlet_sheets, resolved, project_name):
    """
    For NG-MRA and Nigeria Ville only:
    if a Data Queries stock/purchase row also appears in Outlet Stock/Outlet Purchases
    by Outlet ID + SKU ID, mark it as ignored using Previous Month Feedback = Outlet Query.
    """
    if final_df is None or final_df.empty:
        return final_df
    if not _project_uses_outlet_query_ignore(project_name):
        return final_df

    oc = resolved.get('outlet_id')
    sc = resolved.get('sku_id')
    stock_col = resolved.get('stock_col')
    purchase_col = resolved.get('purchase_col')
    if not oc or not sc or not stock_col or not purchase_col:
        return final_df

    outlet_keys = set()
    for sheet_name in ['Outlet Stock', 'Outlet Purchases']:
        odf = outlet_sheets.get(sheet_name, pd.DataFrame()) if isinstance(outlet_sheets, dict) else pd.DataFrame()
        if odf is None or odf.empty or oc not in odf.columns or sc not in odf.columns:
            continue
        for _, r in odf[[oc, sc]].dropna(how='any').iterrows():
            outlet_keys.add(str(r[oc]).strip() + str(r[sc]).strip())

    if not outlet_keys:
        return final_df

    df = final_df.copy()
    PF = 'Previous Month Feedback'
    if PF not in df.columns:
        df[PF] = np.nan

    stock_norm = normalize(stock_col)
    purchase_norm = normalize(purchase_col)
    price_norms = {normalize(c) for c in (resolved.get('price_cols') or [])}
    if resolved.get('buying_price_col'):
        price_norms.add(normalize(resolved.get('buying_price_col')))

    ignored = 0
    for idx, row in df.iterrows():
        if oc not in row.index or sc not in row.index:
            continue
        key = str(row.get(oc, '')).strip() + str(row.get(sc, '')).strip()
        if key not in outlet_keys:
            continue

        q_parts = [q.strip() for q in str(row.get('Queries', '')).split(',') if q.strip()]
        q_norms = {normalize(q) for q in q_parts}

        has_stock_purchase = bool(q_norms & {stock_norm, purchase_norm})
        has_other_data_query = bool(q_norms - {stock_norm, purchase_norm})

        # Do not hide mixed rows that also contain price/profit/facings/negative sales.
        # Only move pure stock/purchase rows to ignored.
        if has_stock_purchase and not has_other_data_query:
            existing = row.get(PF)
            existing_txt = '' if pd.isna(existing) else str(existing).strip()
            df.at[idx, PF] = 'Outlet Query' if not existing_txt else existing_txt + ' | Outlet Query'
            ignored += 1

    if ignored:
        print(f"   🚫 {ignored} stock/purchase Data Queries moved to Ignored because they are Outlet Queries")
    return df


def _derive_output_paths(data_path, output_dir):
    # output_dir is now only the folder where the two output files will be saved.
    # File names are automatically generated from the input file name.
    out_dir = output_dir or os.path.dirname(data_path) or '.'
    base = os.path.splitext(os.path.basename(data_path))[0]
    return (
        os.path.join(out_dir, f'{base}-Data Queries.xlsx'),
        os.path.join(out_dir, f'{base}-General Queries.xlsx')
    )


def _export_data_queries_only(df, output_file, non_flagged_dict=None):
    """Same current Data Queries workbook, but without Final Queries sheet."""
    PF = 'Previous Month Feedback'
    has_pf = PF in df.columns
    if has_pf:
        df_ign = df[df[PF].notna() & (df[PF].astype(str).str.strip()!='')].copy()
        df_act = df[df[PF].isna()  | (df[PF].astype(str).str.strip()=='')].copy()
    else:
        df_ign = pd.DataFrame(columns=df.columns)
        df_act = df.copy()

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        wb = writer.book
        df.to_excel(writer, sheet_name='All Queries', index=False)
        _write_sheet(wb, writer.sheets['All Queries'], df, grey_rows=True)

        if not df_ign.empty:
            df_ign.to_excel(writer, sheet_name='Ignored Queries', index=False)
            _write_sheet(wb, writer.sheets['Ignored Queries'], df_ign, all_grey=True)
        else:
            ws = wb.add_worksheet('Ignored Queries')
            ws.write(0, 0, 'No ignored queries.', wb.add_format({'italic':True,'font_color':'#808080'}))

        df_act.to_excel(writer, sheet_name='Active Queries', index=False)
        _write_sheet(wb, writer.sheets['Active Queries'], df_act, grey_rows=False)

        if non_flagged_dict:
            for metric_col, nf_df in non_flagged_dict.items():
                if nf_df is None or nf_df.empty:
                    continue
                sheet_name = f'NF: {metric_col}'[:31]
                for ch in ['/', '\\', '?', '*', '[', ']', ':']:
                    sheet_name = sheet_name.replace(ch, ' ')
                try:
                    nf_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    _write_sheet(wb, writer.sheets[sheet_name], nf_df, grey_rows=False)
                except Exception as e:
                    print(f"   WARNING: Could not write sheet '{sheet_name}': {e}")

    print(f'   Data Queries file: {output_file}')
    print(f'   Sheet "All Queries":     {len(df):,} rows')
    print(f'   Sheet "Ignored Queries": {len(df_ign):,} rows')
    print(f'   Sheet "Active Queries":  {len(df_act):,} rows')


def _excel_safe_sheet_name(name):
    name = str(name)[:31]
    for ch in ['/', '\\', '?', '*', '[', ']', ':']:
        name = name.replace(ch, ' ')
    return name



def _merge_repeated_outlet_blocks(ws, merge_through_col, outletid_col=1, header_row=1):
    """Merge repeated outlet-level columns for Outlet Stock/Purchases sheets."""
    from openpyxl.styles import Alignment
    max_row = ws.max_row
    if max_row <= header_row + 1 or merge_through_col < outletid_col:
        return

    align = Alignment(vertical="center", horizontal="center", wrap_text=False)
    start = header_row + 1
    current = ws.cell(row=start, column=outletid_col).value

    for r in range(start + 1, max_row + 1):
        val = ws.cell(row=r, column=outletid_col).value
        if val != current:
            end = r - 1
            if end > start:
                for c in range(outletid_col, merge_through_col + 1):
                    ws.merge_cells(start_row=start, start_column=c, end_row=end, end_column=c)
                    ws.cell(row=start, column=c).alignment = align
            start = r
            current = val

    end = max_row
    if end > start:
        for c in range(outletid_col, merge_through_col + 1):
            ws.merge_cells(start_row=start, start_column=c, end_row=end, end_column=c)
            ws.cell(row=start, column=c).alignment = align

def _export_queries_workbook(output_file, sheets_dict):
    """Second workbook: Data Queries + Outlet Stock + Outlet Purchases + New/Lost + SKU Counts."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            safe_name = _excel_safe_sheet_name(sheet_name)
            if df is None or df.empty:
                pd.DataFrame({'Message': [f'No records for {sheet_name}.']}).to_excel(writer, sheet_name=safe_name, index=False)
            else:
                df.to_excel(writer, sheet_name=safe_name, index=False)

        wb = writer.book
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D9D9D9')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            is_data_queries_sheet = ws.title == 'Data Queries'

            # Header style. Do not wrap Data Queries sheet so it stays readable.
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=False if is_data_queries_sheet else True
                )

            # Body style. Data Queries should not have wrapped cells.
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical='top',
                        wrap_text=False if is_data_queries_sheet else True
                    )

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    val = '' if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val), 45))
                ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 45))

            # Match the original Outlets output: merge repeated outlet-level blocks
            # across outlet details and outlet-month-sum columns, leaving SKU rows separate.
            if ws.title in ('Outlet Stock', 'Outlet Purchases') and ws.max_row > 2:
                headers = [cell.value for cell in ws[1]]
                sku_header_idx = None
                for i, h in enumerate(headers, start=1):
                    if normalize(h) in ('prodcode', 'whskuid', 'skuid', 'sku', 'productcode'):
                        sku_header_idx = i
                        break
                if sku_header_idx and sku_header_idx > 1:
                    _merge_repeated_outlet_blocks(ws, merge_through_col=sku_header_idx - 1)

    print(f'   General Queries file: {output_file}')
    for name, df in sheets_dict.items():
        print(f'   Sheet "{name}": {0 if df is None else len(df):,} rows')


# ═══════════════════════════════════════════════════════════════
# PART 11: MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def run_data_queries(project_name, data_path, output_dir, prev_feedback_path=None):
    t0 = time.time()
    print('\n' + '='*60)
    print(f'  DATA QUERY ENGINE — {project_name}')
    print('='*60)

    if project_name not in PROJECT_CONFIGS:
        print(f"ERROR: Unknown project '{project_name}'")
        print(f"   Available: {list(PROJECT_CONFIGS.keys())}")
        return

    config = PROJECT_CONFIGS[project_name]
    df_list, all_months, curr_m, prev_m, sheet_names = load_sheets(data_path)

    print('\n🔍 Resolving columns...')
    resolved = resolve_all(df_list[0], config)

    units_map = {}
    if config.get('use_units_in_price_comb'):
        print('🔢 Resolving units...')
        units_map = build_units_map(df_list, all_months, resolved)

    _crit = {'outlet_id': resolved.get('outlet_id'),
             'sku_id':    resolved.get('sku_id')}
    _miss = [k for k,v in _crit.items() if v is None]
    if _miss:
        print(f"\nERROR: Critical columns not found: {_miss}")
        print(f"   PROJECT_NAME is: '{project_name}'")
        print(f"   File columns: {list(df_list[0].columns[:10])}")
        return

    df_list = [add_comb_keys(df, resolved, config, units_map) for df in df_list]

    price_cols  = resolved.get('price_cols', [])
    if resolved.get('buying_price_col'):
        price_cols = price_cols + [config['buying_price_col']]
    stock_cols  = [c for c in [resolved.get('stock_col'),
                                resolved.get('purchase_col')] if c]
    all_metric_cols = price_cols + stock_cols

    print('\n🔎 Special checks...')
    special_rows = []
    if any(c in config.get('extra_checks',[]) for c in ['negative_profit','no_profit']):
        special_rows.extend(check_profit(df_list[-1], resolved, df_list, all_months, config))
    facings_rows = check_facings(df_list[-1], resolved)
    special_rows.extend(facings_rows)

    print('\n📉 Calculating sales...')
    neg_rows, sales_dict = calculate_sales(df_list, all_months, resolved, all_metric_cols)

    final_rows = []
    non_flagged = {}

    if price_cols:
        print(f'\n📊 Comb IQR — Price: {price_cols}')
        p_res, p_nf = run_comb_detection(
            df_list, all_months, sheet_names, curr_m, prev_m,
            price_cols, resolved, config, comb_type='price')
        final_rows.extend(build_comb_rows(
            p_res, df_list, all_months, curr_m, resolved,
            price_cols, sales_dict, comb_type='price'))
        non_flagged.update(p_nf)
        print(f'   Comb Price: {sum(1 for r in final_rows if isinstance(r,dict))} flagged')

    if stock_cols:
        print(f'\n📊 Comb IQR — Stock/Purchase: {stock_cols}')
        s_res, s_nf = run_comb_detection(
            df_list, all_months, sheet_names, curr_m, prev_m,
            stock_cols, resolved, config, comb_type='stock')
        stock_rows = build_comb_rows(
            s_res, df_list, all_months, curr_m, resolved,
            stock_cols, sales_dict, comb_type='stock')
        final_rows.extend(stock_rows)
        non_flagged.update(s_nf)
        print(f'   Comb Stock: {len(stock_rows)} flagged')

    if price_cols:
        print('\n📊 Comb2 IQR — Price')
        final_rows.extend(run_comb2_detection(
            df_list, all_months, curr_m, price_cols, resolved, config))

    all_rows = neg_rows + special_rows + [r for r in final_rows if isinstance(r, dict)]
    if not all_rows:
        print('\nWARNING: No outliers detected.')
        return

    combined_df = pd.DataFrame(all_rows)
    combined_df['Comb']    = combined_df['Comb'].astype(str)
    combined_df['Queries'] = combined_df['Queries'].astype(str)

    def _collapse(x):
        if x.name in ['Queries','Combination']:
            items = []
            for val in x.dropna():
                items.extend(p.strip() for p in str(val).split(',') if p.strip())
            return ', '.join(sorted(set(items)))
        nn = x.dropna()
        return nn.iloc[0] if len(nn)>0 else np.nan

    final_df = combined_df.groupby('Comb', as_index=False).agg(_collapse)

    if prev_feedback_path:
        sup = load_suppression_set(prev_feedback_path, resolved)
        final_df = apply_suppression(final_df, sup, resolved)

    final_df = apply_new_baseline_check(
        final_df, df_list, all_months, resolved, stock_cols)

    print('\n🔄 Checking spike-and-return...')
    final_df = apply_spike_and_return_check(
        final_df, df_list, all_months, resolved, all_metric_cols, price_cols)

    final_df = reorder_output(final_df, all_months, all_metric_cols)

    # Build outlet sheets before exporting so NG-MRA/Nigeria Ville can move duplicated
    # stock/purchase Data Queries into Ignored Queries as Outlet Query.
    outlet_sheets = build_outlet_query_sheets(df_list, all_months, resolved)
    final_df = _apply_outlet_query_ignore(final_df, outlet_sheets, resolved, project_name)

    data_queries_output, queries_output = _derive_output_paths(data_path, output_dir)
    os.makedirs(os.path.dirname(data_queries_output), exist_ok=True)

    # First workbook: existing Data Queries output, without Final Queries sheet.
    _export_data_queries_only(final_df, data_queries_output, non_flagged_dict=non_flagged)

    # Second workbook: combined field-facing Queries output.
    PF = 'Previous Month Feedback'
    if PF in final_df.columns:
        df_active_for_queries = final_df[final_df[PF].isna() | (final_df[PF].astype(str).str.strip()=='')].copy()
    else:
        df_active_for_queries = final_df.copy()

    data_queries_sheet = build_final_queries(df_active_for_queries, all_months, resolved)
    new_lost_sheet = build_new_lost_products(df_list, all_months, resolved)
    sku_counts_sheet = build_sku_counts(df_list, all_months, resolved)

    combined_query_sheets = {
        'Data Queries': data_queries_sheet,
        'Outlet Stock': outlet_sheets.get('Outlet Stock', pd.DataFrame()),
        'Outlet Purchases': outlet_sheets.get('Outlet Purchases', pd.DataFrame()),
        'New and Lost Products': new_lost_sheet,
        'SKU Counts': sku_counts_sheet,
    }
    _export_queries_workbook(queries_output, combined_query_sheets)

    PF      = 'Previous Month Feedback'
    n_ign   = (final_df[PF].notna()&(final_df[PF].astype(str).str.strip()!='')).sum() \
              if PF in final_df.columns else 0
    elapsed = time.time() - t0
    print(f'\n{"="*60}')
    print(f'  SUMMARY — {project_name} | {curr_m}')
    print(f'{"="*60}')
    print(f'  Total flagged:      {len(final_df):,}')
    print(f'  → Active:           {len(final_df)-n_ign:,}')
    print(f'  → Ignored:          {n_ign:,}')
    print(f'  Negative Sales:     {len(neg_rows):,}')
    print(f'  Facings:            {len(facings_rows):,}')
    if special_rows:
        print(f'  Profit:             {len(special_rows)-len(facings_rows):,}')
    print(f'  Non-Flagged sheets: {len(non_flagged)}')
    print(f'  ⏱  Runtime:         {elapsed:.1f}s')
    print(f'\n  Saved to folder:\n   {output_dir}')
    print(f'  Data Queries file:\n   {data_queries_output}')
    print(f'  General Queries file:\n   {queries_output}')


if __name__ == '__main__':
    run_data_queries(PROJECT_NAME, DATA_PATH, OUTPUT_DIR, PREV_FEEDBACK_PATH)
